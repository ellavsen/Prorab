import asyncio
import html
import logging
import os
import re
import io
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, Tuple, Dict, List

from dotenv import load_dotenv
load_dotenv()

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
)
from telegram.constants import ParseMode
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
    CallbackQueryHandler,
)

from sqlalchemy import (
    create_engine,
    Integer,
    String,
    DateTime,
    select,
    delete,
    func,
    text,
    ForeignKey,
    inspect as sa_inspect,
)
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker, Session

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from smeta_core import (
    Category,
    EstimateTotals,
    PositionData,
    calculate_estimate,
    format_money,
    format_qty,
    from_bp,
    from_kop,
    from_milli,
    merge_duplicates,
    parse_position_line,
    parse_price,
    parse_quantity,
    to_bp,
    to_kop,
    to_milli,
)

# =========================
# Config & Logging
# =========================

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("estimate-bot")

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TOKEN:
    raise RuntimeError("Не задан TELEGRAM_BOT_TOKEN в окружении. Проверь .env")

DB_URL = os.getenv("ESTIMATE_DB_URL", "sqlite:///estimate.db")
# Наценка по умолчанию для НОВЫХ смет, в базисных пунктах: 600 = 6.00%.
# Ставка копируется в смету при создании и дальше живёт в документе (ADR-002).
DEFAULT_MARKUP_BP = 600
RETENTION_LIMIT = 5          # хранить последние N смет на пользователя

# =========================
# Database setup
# =========================

class Base(DeclarativeBase):
    pass

class Estimate(Base):
    __tablename__ = "estimates"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    user_id: Mapped[int] = mapped_column(Integer, index=True)
    number: Mapped[int] = mapped_column(Integer)  # нумерация внутри пользователя
    name: Mapped[str] = mapped_column(String(255))
    # Наценка — часть документа, а не глобальная настройка (ADR-002).
    markup_work_bp: Mapped[int] = mapped_column(Integer, default=DEFAULT_MARKUP_BP)
    markup_material_bp: Mapped[int] = mapped_column(Integer, default=DEFAULT_MARKUP_BP)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    @property
    def markup_work_rate(self) -> Decimal:
        return from_bp(self.markup_work_bp)

    @property
    def markup_material_rate(self) -> Decimal:
        return from_bp(self.markup_material_bp)

class Position(Base):
    __tablename__ = "positions"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    user_id: Mapped[int] = mapped_column(Integer, index=True)
    estimate_id: Mapped[Optional[int]] = mapped_column(ForeignKey("estimates.id"), nullable=True)
    category: Mapped[str] = mapped_column(String(16))  # "Материал" | "Работа"
    name: Mapped[str] = mapped_column(String(255))
    unit: Mapped[str] = mapped_column(String(32), default="")
    # Целые минорные единицы: SQLite хранит NUMERIC как REAL, то есть деньги
    # в Numeric(18,2) уже лежали бы в binary float (ADR-004).
    qty_milli: Mapped[int] = mapped_column(Integer, default=0)
    price_kop: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

    @property
    def qty(self) -> Decimal:
        return from_milli(self.qty_milli)

    @property
    def price(self) -> Decimal:
        return from_kop(self.price_kop)

    def to_domain(self) -> PositionData:
        return PositionData(
            category=Category(self.category),
            name=self.name,
            qty=self.qty,
            price=self.price,
            unit=self.unit or "",
        )

engine = create_engine(DB_URL, future=True)
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False, future=True)

def migrate_money_to_integers(conn) -> List[str]:
    """Переводит qty/price из REAL в целые минорные единицы (ADR-004).

    Возвращает журнал строк, чьё хранимое значение не совпало с копейками
    после округления — это и есть накопленная погрешность REAL-хранения.
    """
    conn.execute(text("ALTER TABLE positions ADD COLUMN qty_milli INTEGER DEFAULT 0"))
    conn.execute(text("ALTER TABLE positions ADD COLUMN price_kop INTEGER DEFAULT 0"))

    journal: List[str] = []
    rows = conn.execute(text("SELECT id, qty, price FROM positions")).fetchall()
    for rid, qty_raw, price_raw in rows:
        qty = Decimal(str(qty_raw if qty_raw is not None else 0))
        price = Decimal(str(price_raw if price_raw is not None else 0))
        qty_milli = int(qty.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP).scaleb(3))
        price_kop = int(price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP).scaleb(2))
        if Decimal(qty_milli).scaleb(-3) != qty or Decimal(price_kop).scaleb(-2) != price:
            journal.append(f"#{rid}: qty {qty} -> {qty_milli}/1000, price {price} -> {price_kop}/100")
        conn.execute(
            text("UPDATE positions SET qty_milli = :q, price_kop = :p WHERE id = :i"),
            {"q": qty_milli, "p": price_kop, "i": rid},
        )

    for column in ("qty", "price"):
        try:
            conn.execute(text(f"ALTER TABLE positions DROP COLUMN {column}"))
        except Exception:
            logger.warning("Старая колонка positions.%s не удалена — SQLite слишком старый", column)
    return journal


def bootstrap_schema_and_migrate():
    """Создаём таблицы и мягко мигрируем старые данные (если нет estimates/estimate_id)."""
    with engine.begin() as conn:
        inspector = sa_inspect(conn)
        tables = inspector.get_table_names()

        # если база пустая — создаём всё сразу
        if "positions" not in tables and "estimates" not in tables:
            Base.metadata.create_all(bind=conn)
            return

        if "estimates" not in tables:
            Base.metadata.create_all(bind=conn, tables=[Estimate.__table__])
        else:
            est_cols = {c["name"] for c in inspector.get_columns("estimates")}
            for column in ("markup_work_bp", "markup_material_bp"):
                if column not in est_cols:
                    conn.execute(text(
                        f"ALTER TABLE estimates ADD COLUMN {column} INTEGER "
                        f"NOT NULL DEFAULT {DEFAULT_MARKUP_BP}"
                    ))

        cols = {c["name"] for c in inspector.get_columns("positions")} if "positions" in tables else set()
        if "positions" not in tables:
            Base.metadata.create_all(bind=conn, tables=[Position.__table__])
            return

        if "estimate_id" not in cols:
            conn.execute(text("ALTER TABLE positions ADD COLUMN estimate_id INTEGER"))

        if "price_kop" not in cols:
            journal = migrate_money_to_integers(conn)
            logger.info("Миграция денег в целые: строк с расхождением — %d", len(journal))
            for entry in journal:
                logger.info("  %s", entry)

        # Создадим дефолтную смету для пользователей, у кого есть позиции без estimate_id
        users = conn.execute(text("SELECT DISTINCT user_id FROM positions WHERE estimate_id IS NULL")).fetchall()
        for (uid,) in users:
            row = conn.execute(text("SELECT MAX(number) FROM estimates WHERE user_id = :uid"), {"uid": uid}).fetchone()
            next_num = (row[0] or 0) + 1
            name = f"Смета №{next_num}"
            now = datetime.utcnow().isoformat(sep=" ")
            conn.execute(
                text("INSERT INTO estimates (user_id, number, name, created_at, updated_at) VALUES (:uid, :num, :name, :c, :u)"),
                {"uid": uid, "num": next_num, "name": name, "c": now, "u": now},
            )
            est_id = conn.execute(
                text("SELECT id FROM estimates WHERE user_id = :uid AND number = :num"),
                {"uid": uid, "num": next_num}
            ).fetchone()[0]
            conn.execute(
                text("UPDATE positions SET estimate_id = :eid WHERE user_id = :uid AND estimate_id IS NULL"),
                {"eid": est_id, "uid": uid}
            )

bootstrap_schema_and_migrate()

# =========================
# Runtime per-user state
# =========================

user_category: Dict[int, Optional[str]] = {}      # "Материал"|"Работа"|None
current_estimate_cache: Dict[int, int] = {}       # user_id -> estimate_id (для быстрого доступа)

# =========================
# Helpers: parsing & formatting
# =========================

def esc(s) -> str:
    """Экранирует пользовательский текст для сообщений с parse_mode=HTML."""
    return html.escape(str(s))

def reply_kb_start():
    return ReplyKeyboardMarkup([[KeyboardButton("Начнём")]], resize_keyboard=True, one_time_keyboard=True)

def reply_kb_categories():
    return ReplyKeyboardMarkup([[KeyboardButton("Работа"), KeyboardButton("Материал")]], resize_keyboard=True)

# Разбор строк ввода живёт в smeta_core.parsing — там же валидация границ.

# =========================
# Estimate helpers
# =========================

def get_current_estimate(db: Session, uid: int) -> Estimate:
    """Берём из кэша или самую свежую смету; если нет — создаём №1."""
    if uid in current_estimate_cache:
        est = db.get(Estimate, current_estimate_cache[uid])
        if est:
            return est

    est = db.execute(
        select(Estimate).where(Estimate.user_id == uid).order_by(Estimate.updated_at.desc())
    ).scalars().first()
    if est:
        current_estimate_cache[uid] = est.id
        return est

    num = 1
    name = f"Смета №{num}"
    est = Estimate(user_id=uid, number=num, name=name)
    db.add(est)
    db.commit()
    db.refresh(est)
    current_estimate_cache[uid] = est.id
    return est

def set_current_estimate(uid: int, est_id: int):
    current_estimate_cache[uid] = est_id

def next_estimate_number(db: Session, uid: int) -> int:
    row = db.execute(select(func.max(Estimate.number)).where(Estimate.user_id == uid)).scalar()
    return (row or 0) + 1

def enforce_retention(db: Session, uid: int):
    """Храним только последние RETENTION_LIMIT смет (по updated_at). Остальные удаляем с позициями."""
    ests: List[Estimate] = db.execute(
        select(Estimate).where(Estimate.user_id == uid).order_by(Estimate.updated_at.desc(), Estimate.id.desc())
    ).scalars().all()
    if len(ests) <= RETENTION_LIMIT:
        return
    to_keep = ests[:RETENTION_LIMIT]
    to_delete = ests[RETENTION_LIMIT:]
    keep_ids = {e.id for e in to_keep}
    del_ids = [e.id for e in to_delete]
    if not del_ids:
        return
    # Удалим позиции этих смет
    db.execute(
        delete(Position).where(Position.user_id == uid, Position.estimate_id.in_(del_ids))
    )
    # Удалим сами сметы
    db.execute(
        delete(Estimate).where(Estimate.user_id == uid, Estimate.id.in_(del_ids))
    )
    db.commit()
    # Если активная смета исчезла — переключим на самую свежую из оставшихся
    if uid in current_estimate_cache and current_estimate_cache[uid] not in keep_ids:
        fresh = to_keep[0]
        current_estimate_cache[uid] = fresh.id

def touch_estimate(db: Session, est: Estimate):
    est.updated_at = datetime.utcnow()
    db.commit()

def load_positions(db: Session, uid: int, est_id: int) -> List[Position]:
    return db.execute(
        select(Position)
        .where(Position.user_id == uid, Position.estimate_id == est_id)
        .order_by(Position.category, Position.id)
    ).scalars().all()

def estimate_totals(db: Session, uid: int, est: Estimate) -> EstimateTotals:
    """Единственный способ узнать сумму сметы — во всех каналах один и тот же."""
    rows = load_positions(db, uid, est.id)
    return calculate_estimate(
        [r.to_domain() for r in rows], est.markup_work_rate, est.markup_material_rate
    )

def markup_caption(est: Estimate) -> str:
    if est.markup_work_bp == est.markup_material_bp:
        return f"{format_money(est.markup_work_rate)}%"
    return (
        f"работы {format_money(est.markup_work_rate)}%, "
        f"материалы {format_money(est.markup_material_rate)}%"
    )

def create_new_estimate_like(db: Session, uid: int, src_est: Estimate) -> Estimate:
    """Создаёт новую пустую смету с тем же названием, но с новым номером, делает её активной и применяет ретеншн."""
    num = next_estimate_number(db, uid)
    new_est = Estimate(
        user_id=uid,
        number=num,
        name=src_est.name,
        # Ставки — часть документа, поэтому переносим из исходной сметы,
        # а не берём текущие настройки: «обновить» значит «то же, но заново».
        markup_work_bp=src_est.markup_work_bp,
        markup_material_bp=src_est.markup_material_bp,
    )
    db.add(new_est)
    db.commit()
    db.refresh(new_est)
    set_current_estimate(uid, new_est.id)
    enforce_retention(db, uid)
    return new_est

# =========================
# Handlers
# =========================

START_TEXT = (
    "Привет! Я бот для расчёта смет.\n\n"
    "📌 Как пользоваться:\n"
    "— Нажми <b>«Начнём»</b>\n"
    "— Выбери категорию: <b>Материал</b> или <b>Работа</b>\n"
    "— Вводи строки:\n"
    "   🪵 Материал: <code>Гвозди, 1000, 20</code>  (кол-во, цена)\n"
    "   👷 Работа:   <code>Побелка, 150, 3000</code> (кол-во, цена)\n\n"
    "Сметы:\n"
    "/new [название] — новая смета и переключение на неё\n"
    "/estimates — список последних 5 смет\n"
    "/switch N — переключиться на смету №N\n\n"
    "Позиции (в рамках текущей сметы):\n"
    "/list — список позиций\n"
    "/delete ID — удалить позицию\n"
    "/edit ID [количество] [цена] — изменить\n"
    "/generate — Excel по текущей смете\n"
    "/clear — очистить позиции текущей сметы (с подтверждением)\n"
)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.effective_message.reply_text(START_TEXT, reply_markup=reply_kb_start(), parse_mode=ParseMode.HTML)

async def handle_begin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip().lower()
    if text == "начнём":
        user_category[update.effective_user.id] = None
        await update.message.reply_text("Выбери категорию: «Работа» или «Материал».", reply_markup=reply_kb_categories())
        return
    await handle_category(update, context)

async def handle_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = (update.message.text or "").strip().lower()
    uid = update.effective_user.id
    if txt in ("работа", "материал"):
        cat = "Работа" if txt == "работа" else "Материал"
        user_category[uid] = cat
        example = "Гвозди, 1000, 20" if cat == "Материал" else "Побелка, 150, 3000"
        await update.message.reply_text(
            f"Активная категория: <b>{cat}</b> ✅\n"
            f"Введи позиции построчно. Пример: <code>{example}</code>\n"
            f"Когда закончишь — /generate",
            parse_mode=ParseMode.HTML,
            reply_markup=reply_kb_categories(),
        )
    else:
        await update.message.reply_text("Нажми «Работа» или «Материал», либо /help.", reply_markup=reply_kb_categories())

# --- Estimates management ---

async def cmd_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    parts = (update.message.text or "").strip().split(" ", 1)
    custom_name = parts[1].strip() if len(parts) > 1 else None

    with SessionLocal() as db:
        num = next_estimate_number(db, uid)
        est_name = custom_name or f"Смета №{num}"
        est = Estimate(user_id=uid, number=num, name=est_name)
        db.add(est)
        db.commit()
        db.refresh(est)
        set_current_estimate(uid, est.id)
        enforce_retention(db, uid)

    await update.message.reply_text(
        f"Создана и активирована <b>Смета №{est.number}</b> — {esc(est.name)}",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_kb_categories(),
    )

async def cmd_estimates(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    with SessionLocal() as db:
        ests = db.execute(
            select(Estimate)
            .where(Estimate.user_id == uid)
            .order_by(Estimate.updated_at.desc(), Estimate.id.desc())
            .limit(RETENTION_LIMIT)
        ).scalars().all()

        if not ests:
            await update.message.reply_text("У тебя пока нет смет. Создай: /new")
            return

        # Отправляем каждую смету отдельным сообщением с кнопкой "Обновить смету №N"
        for e in ests:
            # Итог считает домен. Денежных агрегатов в SQL нет — они дают
            # другой ответ, чем /list и Excel (ADR-002).
            totals = estimate_totals(db, uid, e)
            count = len(totals.lines)
            is_active = (uid in current_estimate_cache and current_estimate_cache[uid] == e.id)
            mark = " (активная)" if is_active else ""
            text_msg = (
                f"№{e.number}: {e.name}{mark}\n"
                f"Позиции: {count}  Итого: {format_money(totals.total)}"
            )
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton(f"Обновить смету №{e.number}", callback_data=f"renew:{e.id}")
            ]])
            await update.message.reply_text(text_msg, reply_markup=kb)

async def cmd_switch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = (update.message.text or "").split()
    if len(args) < 2:
        await update.message.reply_text("Использование: /switch N (номер сметы)")
        return
    try:
        num = int(args[1])
    except Exception:
        await update.message.reply_text("Номер сметы должен быть числом. Пример: /switch 3")
        return
    uid = update.effective_user.id
    with SessionLocal() as db:
        est = db.execute(select(Estimate).where(Estimate.user_id == uid, Estimate.number == num)).scalars().first()
        if not est:
            await update.message.reply_text(f"Смета №{num} не найдена. Посмотри /estimates")
            return
        set_current_estimate(uid, est.id)
        touch_estimate(db, est)
    await update.message.reply_text(
        f"Переключился на <b>Смета №{num}</b> — {esc(est.name)}",
        parse_mode=ParseMode.HTML,
        reply_markup=reply_kb_categories(),
    )

# --- Positions ---

async def add_line(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    cat = user_category.get(uid)
    if not cat:
        await update.message.reply_text("Сначала выбери категорию: «Работа» или «Материал».", reply_markup=reply_kb_categories())
        return
    lines = [l.strip() for l in (update.message.text or "").split("\n")]
    lines = [l for l in lines if l]

    added = 0
    with SessionLocal() as db:
        est = get_current_estimate(db, uid)
        for line in lines:
            try:
                position = parse_position_line(line, Category(cat))

                # дубликаты: (estimate_id, category, name, price)
                existing = db.execute(
                    select(Position).where(
                        Position.user_id == uid,
                        Position.estimate_id == est.id,
                        Position.category == cat,
                        Position.name == position.name,
                        Position.price_kop == to_kop(position.price),
                    )
                ).scalars().first()
                if existing:
                    merged = merge_duplicates([existing.to_domain(), position])[0]
                    existing.qty_milli = to_milli(merged.qty)
                else:
                    db.add(Position(
                        user_id=uid,
                        estimate_id=est.id,
                        category=cat,
                        name=position.name,
                        qty_milli=to_milli(position.qty),
                        price_kop=to_kop(position.price),
                        unit="",
                    ))
                added += 1
            except Exception as e:
                await update.message.reply_text(
                    f"❗️Не удалось добавить строку:\n<code>{esc(line)}</code>\nПричина: {esc(e)}",
                    parse_mode=ParseMode.HTML
                )
        touch_estimate(db, est)

    await update.message.reply_text(f"Добавлено позиций: {added}. Посмотреть список: /list")

async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    with SessionLocal() as db:
        est = get_current_estimate(db, uid)
        rows = load_positions(db, uid, est.id)
        totals = calculate_estimate(
            [r.to_domain() for r in rows], est.markup_work_rate, est.markup_material_rate
        )

    if not rows:
        await update.message.reply_text("В текущей смете пока пусто. Выбери категорию и добавь позиции.", reply_markup=reply_kb_categories())
        return

    out = [f"<b>{esc(est.name)}</b> (№{est.number})"]
    current_cat = None
    for r, line in zip(rows, totals.lines):
        if r.category != current_cat:
            current_cat = r.category
            out.append(f"\n<b>{'Материалы и расходники' if current_cat=='Материал' else 'Работы'}</b>")
        out.append(
            f"#{r.id}: {esc(r.name)}\n"
            f"    Кол-во: {format_qty(r.qty)}  Цена: {format_money(r.price)}  "
            f"Сумма: {format_money(line.total)}"
        )
    out.append(f"\nБез наценки: {format_money(totals.subtotal)}")
    out.append(f"Наценка ({markup_caption(est)}): {format_money(totals.markup)}")
    out.append(f"Итого: <b>{format_money(totals.total)}</b>")
    out.append(f"Наименований: <b>{len(rows)}</b>")

    await update.message.reply_text("\n".join(out), parse_mode=ParseMode.HTML)

async def cmd_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = (update.message.text or "").split()
    if len(args) < 2:
        await update.message.reply_text("Использование: /delete ID")
        return
    try:
        rid = int(args[1])
    except Exception:
        await update.message.reply_text("ID должен быть числом. Пример: /delete 12")
        return

    uid = update.effective_user.id
    with SessionLocal() as db:
        est = get_current_estimate(db, uid)
        row = db.execute(select(Position).where(Position.id == rid, Position.user_id == uid, Position.estimate_id == est.id)).scalars().first()
        if not row:
            await update.message.reply_text("Позиция не найдена в текущей смете.")
            return
        db.delete(row)
        touch_estimate(db, est)
    await update.message.reply_text(f"Удалено: #{rid}. Обновить список: /list")

async def cmd_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    m = re.match(r"^/edit\s+(\d+)(.*)$", text, flags=re.IGNORECASE)
    if not m:
        await update.message.reply_text("Использование: /edit ID [количество] [цена]. Пример: /edit 12 200 350")
        return
    rid = int(m.group(1))
    tail = m.group(2).strip()

    qty: Optional[Decimal] = None
    price: Optional[Decimal] = None

    parts = [p for p in re.split(r"[, ]+", tail) if p != ""]
    try:
        if len(parts) == 1:
            qty = parse_quantity(parts[0])
        elif len(parts) >= 2:
            if parts[0] != "-":
                qty = parse_quantity(parts[0])
            if parts[1] != "-":
                price = parse_price(parts[1])
    except ValueError as e:
        await update.message.reply_text(
            f"{esc(e)}\nПример: <code>/edit 12 200 350</code>", parse_mode=ParseMode.HTML
        )
        return

    uid = update.effective_user.id
    with SessionLocal() as db:
        est = get_current_estimate(db, uid)
        row = db.execute(select(Position).where(Position.id == rid, Position.user_id == uid, Position.estimate_id == est.id)).scalars().first()
        if not row:
            await update.message.reply_text("Позиция не найдена в текущей смете.")
            return
        if qty is not None:
            row.qty_milli = to_milli(qty)
        if price is not None:
            row.price_kop = to_kop(price)

        # Merge duplicates внутри текущей сметы
        dup = db.execute(
            select(Position).where(
                Position.id != row.id,
                Position.user_id == uid,
                Position.estimate_id == est.id,
                Position.category == row.category,
                Position.name == row.name,
                Position.price_kop == row.price_kop,
            )
        ).scalars().first()
        if dup:
            try:
                merged = merge_duplicates([row.to_domain(), dup.to_domain()])[0]
            except ValueError as e:
                await update.message.reply_text(str(e))
                return
            row.qty_milli = to_milli(merged.qty)
            db.delete(dup)

        touch_estimate(db, est)

    await update.message.reply_text(f"Обновлено: #{rid}. Смотри /list")

# ======== /clear с подтверждением ========

async def cmd_clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    with SessionLocal() as db:
        est = get_current_estimate(db, uid)
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("Да", callback_data=f"clear_yes:{est.id}"),
            InlineKeyboardButton("Нет", callback_data=f"clear_no:{est.id}"),
        ]
    ])
    await update.message.reply_text(
        f"Вы точно хотите очистить позиции сметы {est.name} (№{est.number})?",
        reply_markup=kb
    )

# ----- Excel generation -----

THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RATE_CELL = "$B$1"
HEADER_ROW = 3
FIRST_DATA_ROW = 4


def build_sheet(ws, title: str, rows: List[PositionData], rate: Decimal, is_work: bool):
    """Лист с живыми формулами по схеме docs/money.md §3.4.

    Округление живёт внутри формул тем же каскадом, что и в домене, а ставка
    вынесена в отдельную ячейку: заказчик видит, из чего сложился итог, и может
    поменять её честно и на виду.
    """
    ws.title = title
    ws.cell(row=1, column=1, value="Наценка, %").font = Font(bold=True)
    ws.cell(row=1, column=2, value=rate)

    headers = [
        "№",
        "Виды работ" if is_work else "Материалы и расходники",
        "Кол-во",
        "Цена за ед.",
        "Сумма без наценки",
        "Сумма с наценкой",
    ]
    for col, title_text in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=title_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for idx, position in enumerate(rows, start=1):
        row_idx = FIRST_DATA_ROW + idx - 1
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=position.name)
        ws.cell(row=row_idx, column=3, value=position.qty)
        ws.cell(row=row_idx, column=4, value=position.price)
        ws.cell(row=row_idx, column=5, value=f"=ROUND(C{row_idx}*D{row_idx},2)")
        ws.cell(row=row_idx, column=6, value=f"=ROUND(E{row_idx}*(1+{RATE_CELL}/100),2)")

        for col in range(1, 7):
            c = ws.cell(row=row_idx, column=col)
            c.border = BORDER
            c.alignment = Alignment(
                horizontal="center" if col != 2 else "left", vertical="center"
            )

    last_data_row = max(ws.max_row, FIRST_DATA_ROW)

    total_row = last_data_row + 1
    ws.cell(row=total_row, column=2, value="Итого")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{FIRST_DATA_ROW}:E{last_data_row})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{FIRST_DATA_ROW}:F{last_data_row})")

    markup_row = total_row + 1
    ws.cell(row=markup_row, column=2, value="в том числе наценка")
    # Наценка — только разностью. Умножением она разошлась бы с суммой строк.
    ws.cell(row=markup_row, column=6, value=f"=F{total_row}-E{total_row}")

    count_row = markup_row + 1
    ws.cell(row=count_row, column=2, value="Наименований")
    ws.cell(row=count_row, column=3, value=f"=COUNTA(B{FIRST_DATA_ROW}:B{last_data_row})")

    for row_idx in (total_row, markup_row, count_row):
        for col in (2, 3, 5, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(bold=True)
            cell.border = BORDER
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")

    for col, width in enumerate((6, 60, 14, 14, 20, 20), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def positions_by_category(db: Session, uid: int, est_id: int) -> Tuple[List[PositionData], List[PositionData]]:
    """Позиции сметы, разложенные по категориям.

    Слияния дублей здесь нет намеренно. Оно меняет итог — round2(q1*p) +
    round2(q2*p) не равно round2((q1+q2)*p) — поэтому склейка выполняется один
    раз при вводе, до расчёта (money.md §5). Если склеивать ещё и здесь, XLSX
    начнёт расходиться с /list ровно так, как расходился раньше.
    """
    domain = [r.to_domain() for r in load_positions(db, uid, est_id)]
    materials = [p for p in domain if p.category == Category.MATERIAL]
    works = [p for p in domain if p.category == Category.WORK]
    return materials, works

async def cmd_generate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    with SessionLocal() as db:
        est = get_current_estimate(db, uid)
        materials, works = positions_by_category(db, uid, est.id)
        if not materials and not works:
            await update.message.reply_text("Нет данных для отчёта в текущей смете. Добавь позиции и повтори /generate.")
            return

        totals = estimate_totals(db, uid, est)
        wb = Workbook()
        build_sheet(wb.active, "Работы", works, est.markup_work_rate, is_work=True)
        build_sheet(
            wb.create_sheet(), "Материалы и расходники", materials,
            est.markup_material_rate, is_work=False,
        )

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    filename = f"estimate_{uid}_no{est.number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    await update.message.reply_document(
        document=InputFile(bio, filename=filename),
        caption=(
            f"Готово: {est.name} (№{est.number}). Две вкладки: «Работы» и «Материалы».\n"
            f"Наценка {markup_caption(est)}, итог {format_money(totals.total)} — "
            f"столько же, сколько в /list."
        ),
    )

# ----- Callback buttons -----

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    data = q.data or ""
    uid = update.effective_user.id

    # Подтверждение обновления сметы
    if data.startswith("renew:"):
        est_id = int(data.split(":")[1])
        with SessionLocal() as db:
            est = db.get(Estimate, est_id)
            if not est or est.user_id != uid:
                await q.edit_message_text("Смета не найдена.")
                return
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("Да", callback_data=f"renew_yes:{est.id}"),
                    InlineKeyboardButton("Нет", callback_data=f"renew_no:{est.id}"),
                ]
            ])
            await q.edit_message_text(
                f"Обновить смету №{est.number} — {esc(est.name)}?\n\n"
                f"Будет создана новая пустая смета со следующим номером, и она станет активной. Старая останется без изменений.",
                reply_markup=kb
            )
        return

    if data.startswith("renew_yes:"):
        est_id = int(data.split(":")[1])
        with SessionLocal() as db:
            src = db.get(Estimate, est_id)
            if not src or src.user_id != uid:
                await q.edit_message_text("Смета не найдена.")
                return
            new_est = create_new_estimate_like(db, uid, src)
            await q.edit_message_text(
                f"✅ Готово. Создана и активирована <b>Смета №{new_est.number}</b> — {esc(new_est.name)}.\n"
                f"Старая смета №{src.number} осталась без изменений.",
                parse_mode=ParseMode.HTML
            )
        return

    if data.startswith("renew_no:"):
        await q.edit_message_text("Отменено.")
        return

    # Подтверждение очистки сметы
    if data.startswith("clear_yes:"):
        est_id = int(data.split(":")[1])
        with SessionLocal() as db:
            est = db.get(Estimate, est_id)
            if not est or est.user_id != uid:
                await q.edit_message_text("Смета не найдена.")
                return
            db.execute(text("DELETE FROM positions WHERE user_id = :uid AND estimate_id = :eid"),
                       {"uid": uid, "eid": est.id})
            touch_estimate(db, est)
        await q.edit_message_text(f"Очищены позиции сметы {est.name} (№{est.number}).")
        return

    if data.startswith("clear_no:"):
        await q.edit_message_text("Отменено.")
        return

    # Удаление позиции из текущей сметы
    if data.startswith("del:"):
        rid = int(data.split(":")[1])
        with SessionLocal() as db:
            est = get_current_estimate(db, uid)
            row = db.execute(
                select(Position).where(Position.id == rid, Position.user_id == uid, Position.estimate_id == est.id)
            ).scalars().first()
            if not row:
                await q.edit_message_text("Позиция не найдена в текущей смете.")
                return
            db.delete(row)
            touch_estimate(db, est)
        await q.edit_message_text(f"Удалено: #{rid}. Обновить список: /list")
        return

    # Подсказка по редактированию
    if data.startswith("edit:"):
        rid = int(data.split(":")[1])
        await q.edit_message_text(
            f"Чтобы изменить: отправь команду\n<code>/edit {rid} НОВОЕ_КОЛ НОВАЯ_ЦЕНА</code>",
            parse_mode=ParseMode.HTML
        )
        return

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(START_TEXT, parse_mode=ParseMode.HTML, reply_markup=reply_kb_start())

# =========================
# App bootstrap
# =========================

def build_app():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))

    # Estimates
    app.add_handler(CommandHandler("new", cmd_new))
    app.add_handler(CommandHandler("estimates", cmd_estimates))
    app.add_handler(CommandHandler("switch", cmd_switch))

    # Positions & files
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("delete", cmd_delete))
    app.add_handler(CommandHandler("edit", cmd_edit))
    app.add_handler(CommandHandler("generate", cmd_generate))
    app.add_handler(CommandHandler("clear", cmd_clear))  # теперь с подтверждением

    # Buttons & categories (регистронезависимо)
    app.add_handler(MessageHandler(filters.Regex(r"(?i)^начнём$"), handle_begin))
    app.add_handler(MessageHandler(filters.Regex(r"(?i)^(работа|материал)$"), handle_category))

    # Free text lines for positions (current estimate)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, add_line))

    # Inline callbacks
    app.add_handler(CallbackQueryHandler(on_callback))

    return app

def main():
    app = build_app()
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
