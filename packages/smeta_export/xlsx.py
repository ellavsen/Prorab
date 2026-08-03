"""XLSX с живыми формулами по схеме docs/money.md §3.4.

Округление живёт внутри формул тем же каскадом, что и в домене, а ставка
вынесена в отдельную ячейку: заказчик видит, из чего сложился итог, и может
поменять её честно и на виду.
"""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from smeta_core import PositionData, RateBase
from smeta_prices import display_unit

THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")

RATE_CELL = "$B$1"
HEADER_ROW = 3
FIRST_DATA_ROW = 4
COLUMN_WIDTHS = (6, 52, 10, 14, 14, 20, 20)

# Подпись ячейки со ставкой и формула строки — пара: одно без другого лжёт.
# Формула для процента от суммы заказчику ДЕЛИТ, буквально повторяя расчёт
# домена; умножение на обратное — другая операция (money.md §3.5, правило D).
RATE_HEADING = {
    RateBase.COST: "Наценка, % к цене",
    RateBase.PRICE: "По договору, % от суммы",
}
LINE_FORMULA = {
    RateBase.COST: "=ROUND(F{row}*(1+{cell}/100),2)",
    RateBase.PRICE: "=ROUND(F{row}/(1-{cell}/100),2)",
}
MARKUP_ROW_LABEL = {
    RateBase.COST: "в том числе наценка",
    RateBase.PRICE: "в том числе по договору",
}


def build_sheet(
    ws,
    title: str,
    rows: list[PositionData],
    rate: Decimal,
    is_work: bool,
    base: str = RateBase.COST,
) -> None:
    ws.title = title
    ws.cell(row=1, column=1, value=RATE_HEADING[RateBase(base)]).font = Font(bold=True)
    ws.cell(row=1, column=2, value=rate)

    headers = [
        "№",
        "Виды работ" if is_work else "Материалы и расходники",
        "Ед.",
        "Кол-во",
        "Цена за ед.",
        "Сумма без наценки",
        "Сумма с наценкой",
    ]
    for column, caption in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=column, value=caption)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for index, position in enumerate(rows, start=1):
        row = FIRST_DATA_ROW + index - 1
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=position.name)
        # Заказчику печатается сказанное («мешок»), а не канон «шт»: документ
        # должен читаться как разговор с прорабом (ADR-015). Падеж при этом
        # приводится к словарному: в колонке «Ед.» слово стоит само по себе,
        # а не согласовано с количеством из соседней колонки.
        ws.cell(row=row, column=3,
                value=display_unit(position.unit, position.unit_spoken) or "—")
        ws.cell(row=row, column=4, value=position.qty)
        ws.cell(row=row, column=5, value=position.price)
        ws.cell(row=row, column=6, value=f"=ROUND(D{row}*E{row},2)")
        ws.cell(row=row, column=7,
                value=LINE_FORMULA[RateBase(base)].format(row=row, cell=RATE_CELL))

        for column in range(1, 8):
            cell = ws.cell(row=row, column=column)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="left" if column == 2 else "center", vertical="center"
            )

    last_data_row = max(ws.max_row, FIRST_DATA_ROW)

    total_row = last_data_row + 1
    ws.cell(row=total_row, column=2, value="Итого")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{FIRST_DATA_ROW}:F{last_data_row})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{FIRST_DATA_ROW}:G{last_data_row})")

    markup_row = total_row + 1
    ws.cell(row=markup_row, column=2, value=MARKUP_ROW_LABEL[RateBase(base)])
    # Надбавка — только разностью. Умножением она разошлась бы с суммой строк, и
    # на смете это расхождение реально: money.md §3.5, «чего деление не сохраняет».
    ws.cell(row=markup_row, column=7, value=f"=G{total_row}-F{total_row}")

    count_row = markup_row + 1
    ws.cell(row=count_row, column=2, value="Наименований")
    ws.cell(row=count_row, column=4, value=f"=COUNTA(B{FIRST_DATA_ROW}:B{last_data_row})")

    for row in (total_row, markup_row, count_row):
        for column in (2, 4, 6, 7):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(bold=True)
            cell.border = BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

    for column, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def build_workbook(
    materials: list[PositionData],
    works: list[PositionData],
    work_rate: Decimal,
    material_rate: Decimal,
    base: str = RateBase.COST,
) -> io.BytesIO:
    workbook = Workbook()
    build_sheet(workbook.active, "Работы", works, work_rate, is_work=True, base=base)
    build_sheet(
        workbook.create_sheet(), "Материалы и расходники", materials,
        material_rate, is_work=False, base=base,
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
