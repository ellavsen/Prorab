"""API — потребитель домена, а не второй вычислитель."""

import io
import json
from decimal import Decimal as D

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app
from smeta_core import Category, PositionData, calculate_estimate
from smeta_export import build_workbook

client = TestClient(app)

ESTIMATE = {
    "positions": [
        {"category": "work", "name": "Побелка", "qty": "1.5", "price": "100.10"},
        {"category": "work", "name": "Стяжка", "qty": "2.5", "price": "100.10"},
        {"category": "material", "name": "Гвозди", "qty": "1000", "price": "0.37"},
    ],
    "markup_work_rate": "6.00",
    "markup_material_rate": "6.00",
}


def domain_totals(payload=ESTIMATE):
    positions = [
        PositionData(
            category=Category(p["category"]),
            name=p["name"],
            qty=D(p["qty"]),
            price=D(p["price"]),
            unit=p.get("unit", "") or ("шт" if p["category"] == "material" else "м²"),
        )
        for p in payload["positions"]
    ]
    return calculate_estimate(
        positions, D(payload["markup_work_rate"]), D(payload["markup_material_rate"])
    )


def test_healthz():
    assert client.get("/healthz").json() == {"status": "ok", "core": "smeta_core"}


def test_units_are_the_domain_reference():
    from smeta_core import UNITS

    assert client.get("/units").json() == list(UNITS)


def test_calculate_matches_the_domain_exactly():
    body = client.post("/calculate", json=ESTIMATE).json()
    expected = domain_totals()

    assert D(body["total"]) == expected.total
    assert D(body["subtotal"]) == expected.subtotal
    assert D(body["markup"]) == expected.markup
    assert [D(line["total"]) for line in body["lines"]] == [
        line.total for line in expected.lines
    ]


def test_total_equals_the_sum_of_returned_lines():
    body = client.post("/calculate", json=ESTIMATE).json()
    assert sum(D(line["total"]) for line in body["lines"]) == D(body["total"])


def test_money_travels_as_strings_never_as_json_numbers():
    """JSON-число это double. Копейки на границе теряться не должны."""
    raw = json.loads(client.post("/calculate", json=ESTIMATE).content)
    for key in ("subtotal", "markup", "total"):
        assert isinstance(raw[key], str), f"{key} уехало числом: {raw[key]!r}"
    for line in raw["lines"]:
        for key in ("qty", "price", "base", "total"):
            assert isinstance(line[key], str), f"{key} уехало числом: {line[key]!r}"


def test_precision_survives_the_round_trip():
    body = client.post(
        "/calculate",
        json={"positions": [
            {"category": "work", "name": "Тысячные", "qty": "0.007", "price": "5.00"},
        ], "markup_work_rate": "0.00", "markup_material_rate": "0.00"},
    ).json()
    assert body["lines"][0]["qty"] == "0.007"
    assert body["total"] == "0.04"


def test_missing_unit_is_filled_by_category():
    body = client.post("/calculate", json=ESTIMATE).json()
    units = {line["name"]: line["unit"] for line in body["lines"]}
    assert units["Побелка"] == "м²"
    assert units["Гвозди"] == "шт"


def test_unit_filling_can_be_switched_off():
    body = client.post("/calculate", json={**ESTIMATE, "fill_missing_units": False}).json()
    assert all(line["unit"] == "" for line in body["lines"])


@pytest.mark.parametrize(
    "qty, price, expected",
    [
        ("0", "10.00", "больше нуля"),
        ("-5", "10.00", "больше нуля"),
        ("1.2345", "10.00", "не более 3 знаков"),
        ("1", "12.345", "не более 2 знаков"),
        ("100000", "10.00", "не больше"),
    ],
)
def test_domain_refusals_become_422_with_a_readable_reason(qty, price, expected):
    response = client.post(
        "/calculate",
        json={"positions": [{"category": "work", "name": "x", "qty": qty, "price": price}]},
    )
    assert response.status_code == 422
    assert expected in response.json()["detail"]


def test_line_ceiling_is_enforced_over_http():
    response = client.post(
        "/calculate",
        json={"positions": [
            {"category": "work", "name": "x", "qty": "99999.999", "price": "9999999.99"},
        ]},
    )
    assert response.status_code == 422
    assert "превышает потолок" in response.json()["detail"]


def test_xlsx_is_byte_identical_to_what_the_bot_produces():
    response = client.post("/xlsx", json=ESTIMATE)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxml")

    expected = domain_totals()
    positions = [line.position for line in expected.lines]
    reference = build_workbook(
        [p for p in positions if p.category == Category.MATERIAL],
        [p for p in positions if p.category == Category.WORK],
        D("6.00"), D("6.00"),
    )

    from_api = load_workbook(io.BytesIO(response.content))
    from_bot = load_workbook(reference)
    assert from_api.sheetnames == from_bot.sheetnames
    for sheet in from_api.sheetnames:
        a = [[c.value for c in row] for row in from_api[sheet].iter_rows()]
        b = [[c.value for c in row] for row in from_bot[sheet].iter_rows()]
        assert a == b, f"лист «{sheet}» разошёлся"


def test_xlsx_refuses_an_estimate_the_calculator_refuses():
    response = client.post(
        "/xlsx",
        json={"positions": [
            {"category": "work", "name": "x", "qty": "99999.999", "price": "9999999.99"},
        ]},
    )
    assert response.status_code == 422


def test_parse_returns_positions_and_reasons_side_by_side():
    body = client.post("/parse", json={
        "category": "work",
        "text": "Побелка, 150 м2, 3000\nкривая строка\nСтяжка, 40.5, 1200",
    }).json()
    assert [p["name"] for p in body["positions"]] == ["Побелка", "Стяжка"]
    assert body["positions"][0]["unit"] == "м²"
    assert len(body["errors"]) == 1
    assert body["errors"][0]["line"] == "кривая строка"


def test_parse_of_an_empty_text():
    body = client.post("/parse", json={"category": "material", "text": "\n  \n"}).json()
    assert body == {"positions": [], "errors": []}


def test_empty_estimate_is_valid():
    body = client.post("/calculate", json={"positions": []}).json()
    assert body["total"] == "0.00"
    assert body["lines"] == []
