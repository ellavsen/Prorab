"""Главный риск Sprint 7: четыре поверхности одного документа разойдутся.

Заказчик получает смету четырьмя способами — сообщением в боте, файлом XLSX,
файлом PDF и страницей по ссылке. Расхождение хоть на копейку между ними — это
спор о деньгах, в котором прораб не сможет объяснить, какая цифра настоящая.

Сверка идёт **с замороженным итогом**, а не поверхностей друг с другом. Разница
принципиальная: четыре одинаково неверных числа тест «все совпали» пропустил
бы. `frozen_total` — то, что было записано в момент отправки, и с ним обязана
сойтись каждая поверхность по отдельности (money.md И3).

XLSX здесь не читается как файл со значениями: в нём живые формулы, и его
пересчитывает настоящий табличный процессор.
"""

from __future__ import annotations

import io
import os
import pathlib
import shutil
import subprocess
from datetime import date
from decimal import Decimal as D

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from conftest import open_storage
from smeta_core import Category, PositionData, RateBase, format_money
from smeta_export import (
    FIRST_DATA_ROW,
    DocumentMeta,
    build_page,
    build_pdf,
    build_workbook,
)
from smeta_storage import (
    create_estimate,
    positions,
    send,
    set_current_estimate,
    verified_totals,
)

ROOT = pathlib.Path(__file__).resolve().parent.parent
UID = 4242
TODAY = date(2026, 8, 3)
WORK_RATE = D("6.00")
MATERIAL_RATE = D("12.50")

# Данные злые намеренно: обе границы 0.005 в каскаде, копеечная строка,
# крупная строка и разные ставки у работ и материалов. На ровных числах
# четыре поверхности сойдутся и при сломанном округлении.
POSITIONS = [
    PositionData(Category.WORK, "Побелка", D("1.5"), D("100.10"), "м²"),
    PositionData(Category.WORK, "Стяжка", D("2.5"), D("100.10"), "м²"),
    PositionData(Category.WORK, "Половинка", D("0.5"), D("0.01"), "шт"),
    PositionData(Category.WORK, "Крупная", D("999.999"), D("5000.00"), "м.п."),
    PositionData(Category.MATERIAL, "Копейка", D("1"), D("0.01"), "шт"),
    PositionData(Category.MATERIAL, "Цемент М500", D("20"), D("380.00"),
                 unit="", unit_spoken="мешков"),
]


# Замороженные итоги на этих данных. Оба основания проверяются одними и теми
# же поверхностями: расхождение между ними — не «другая цифра», а другой способ
# считать, и разойтись он может ровно там же.
FROZEN = {RateBase.COST: D("5308969.15"), RateBase.PRICE: D("5328255.30")}


@pytest.fixture(params=list(FROZEN), ids=lambda base: str(base))
def document(request, tmp_path):
    """Отправленная смета: у неё есть замороженный итог, с которым сверяемся."""
    base = request.param
    os.environ["ESTIMATE_DB_URL"] = f"sqlite:///{tmp_path / 'parity.db'}"
    _engine, Session = open_storage(tmp_path / "parity.db")
    with Session() as db:
        estimate = create_estimate(
            db, UID, name="Ремонт на объекте",
            markup_work_bp=600, markup_material_bp=1250, rate_base=base,
        )
        set_current_estimate(db, UID, estimate.id)
        for item in POSITIONS:
            positions.add(db, UID, estimate.id, item)
        db.commit()
        send(db, estimate)

        rows = positions.load(db, UID, estimate.id)
        return {
            "base": base,
            "frozen_total": estimate.frozen_total,
            "totals": verified_totals(db, estimate),
            "rows": rows,
            "estimate": estimate,
            "materials_works": positions.by_category(db, UID, estimate.id),
        }


def meta_of(document) -> DocumentMeta:
    estimate = document["estimate"]
    return DocumentMeta(
        number=estimate.number, version=estimate.version, title=estimate.name,
        on=TODAY, work_rate=estimate.markup_work_rate,
        material_rate=estimate.markup_material_rate, status="Отправлена заказчику",
        rate_base=estimate.rate_base,
    )


def test_the_frozen_total_is_what_the_domain_computes(document):
    """Опора всей сверки. Если она сдвинулась, остальные проверки бессмысленны."""
    assert document["totals"].total == document["frozen_total"]
    assert document["frozen_total"] == FROZEN[document["base"]]


# --- Поверхность 1: сообщение в боте ---


def test_the_bot_message_shows_the_frozen_total(document):
    from bot.texts import render_estimate

    text = render_estimate(document["estimate"], document["rows"], document["totals"])
    assert format_money(document["frozen_total"]) in text


# --- Поверхность 2: PDF ---


def test_the_pdf_shows_the_frozen_total(document):
    buffer = build_pdf(document["totals"], meta_of(document))
    text = "\n".join(
        page.extract_text() for page in PdfReader(io.BytesIO(buffer.getvalue())).pages
    ).replace("\xa0", " ")
    assert format_money(document["frozen_total"]) in text


# --- Поверхность 3: публичная страница ---


def test_the_public_page_shows_the_frozen_total(document):
    page = build_page(document["totals"], meta_of(document))
    assert format_money(document["frozen_total"]) in page


# --- Поверхность 4: XLSX, пересчитанный настоящим движком ---


def _libreoffice_missing() -> bool:
    """Пропуск в CI — это «не проверено», а не «зелено» (как в test_xlsx.py)."""
    if shutil.which("soffice") is not None:
        return False
    if os.getenv("REQUIRE_LIBREOFFICE") == "1":
        pytest.fail("REQUIRE_LIBREOFFICE=1, но soffice не найден — паритет не проверен")
    return True


def test_the_workbook_carries_the_same_operation_the_domain_used(document):
    """Формула в ячейке — не иллюстрация расчёта, а он сам.

    Проверяется отдельно от пересчёта, потому что пересчёт требует настоящего
    движка и в его отсутствие пропускается, а подмена деления умножением
    молчаливо разошлась бы с доменом на всех строках сразу.
    """
    materials, works = document["materials_works"]
    book = load_workbook(io.BytesIO(build_workbook(
        materials, works, WORK_RATE, MATERIAL_RATE, document["base"]
    ).getvalue()))
    formula = book["Работы"][f"G{FIRST_DATA_ROW}"].value

    if document["base"] == RateBase.PRICE:
        assert formula == f"=ROUND(F{FIRST_DATA_ROW}/(1-$B$1/100),2)"
    else:
        assert formula == f"=ROUND(F{FIRST_DATA_ROW}*(1+$B$1/100),2)"


@pytest.mark.skipif(_libreoffice_missing(), reason="LibreOffice не установлен")
def test_the_recalculated_workbook_adds_up_to_the_frozen_total(document, tmp_path):
    """Две вкладки, две ставки, живые формулы — и тот же итог до копейки.

    Читаются не записанные значения, а результат пересчёта: в файле лежат
    формулы, и проверять надо то, что увидит заказчик, открыв его.
    """
    materials, works = document["materials_works"]
    source = tmp_path / "estimate.xlsx"
    source.write_bytes(
        build_workbook(
            materials, works, WORK_RATE, MATERIAL_RATE, document["base"]
        ).getvalue()
    )

    out = tmp_path / "out"
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(out), str(source)],
        check=True, capture_output=True, timeout=300,
    )
    book = load_workbook(out / source.name, data_only=True)

    total = D("0.00")
    for name, count in (("Работы", len(works)), ("Материалы и расходники", len(materials))):
        sheet = book[name]
        total_row = FIRST_DATA_ROW + max(count, 1)
        total += D(str(sheet[f"G{total_row}"].value))

    assert total == document["frozen_total"]


# --- Одни и те же слова про процент на всех поверхностях ---

# Что человек обязан прочитать при каждом основании. Проверяется не «есть
# слово», а «нет слова другого основания»: перепутанная подпись опаснее
# отсутствующей — она называет удержание наценкой, и 6% читаются как ×1,06.
WORDING = {
    RateBase.COST: ("наценка", "к цене", "по договору", "от суммы"),
    RateBase.PRICE: ("по договору", "от суммы", "наценка", "к цене"),
}


def test_all_surfaces_name_the_base_of_the_percent_the_same_way(document):
    """Четыре поверхности — один словарь (ADR-024 §5).

    Раньше подпись собирала каждая сама, и разойтись им ничто не мешало: суммы
    сверяются тестами, слова — нет. Теперь слова живут в ядре рядом с самим
    основанием, а здесь проверяется, что до бумаги они доезжают.
    """
    from bot.texts import render_estimate

    word, of, foreign_word, foreign_of = WORDING[document["base"]]
    materials, works = document["materials_works"]

    surfaces = {
        "бот": render_estimate(
            document["estimate"], document["rows"], document["totals"]
        ).lower(),
        "страница": build_page(document["totals"], meta_of(document)).lower(),
        "pdf": "\n".join(
            page.extract_text()
            for page in PdfReader(io.BytesIO(
                build_pdf(document["totals"], meta_of(document)).getvalue()
            )).pages
        ).replace("\xa0", " ").lower(),
        "xlsx": " ".join(
            str(cell.value)
            for sheet in load_workbook(io.BytesIO(build_workbook(
                materials, works, WORK_RATE, MATERIAL_RATE, document["base"]
            ).getvalue()))
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ).lower(),
    }

    for name, text in surfaces.items():
        assert word in text, f"{name}: не сказано «{word}»"
        assert of in text, f"{name}: не сказано, от чего процент («{of}»)"
        assert foreign_word not in text, f"{name}: слово чужого основания «{foreign_word}»"
        assert foreign_of not in text, f"{name}: основание чужое — «{foreign_of}»"


# --- Одно правило форматирования на все текстовые поверхности ---

# Где деньги превращаются в текст для человека. XLSX сюда не входит намеренно:
# он кладёт в ячейку Decimal и формат числа, а печатает его сам Excel — его
# паритет доказывается пересчётом выше, а не сравнением строк.
TEXT_SURFACES = (
    "packages/smeta_export/page.py",
    "packages/smeta_export/pdf.py",
    "apps/bot/texts.py",
    "apps/bot/preview_texts.py",
    "apps/bot/handlers/share.py",
)

# Всё, чем можно отформатировать деньги мимо format_money. Каждое из этого
# даёт точку вместо запятой либо теряет второй знак — то есть ровно то
# расхождение, ради которого написан этот файл.
HOMEBREW = (":.2f", ":,.2f", "%.2f", ":.02f", "round(")


@pytest.mark.parametrize("relative", TEXT_SURFACES)
def test_every_text_surface_formats_money_the_one_way(relative):
    source = (ROOT / relative).read_text(encoding="utf-8")
    assert "format_money" in source, f"{relative}: деньги форматируются мимо домена"
    found = [pattern for pattern in HOMEBREW if pattern in source]
    assert not found, f"{relative}: самодельное форматирование {found}"


def test_the_one_way_is_a_comma_and_two_digits():
    """Формат назван здесь, чтобы его нельзя было поменять незамеченным."""
    assert format_money(D("5308969.15")) == "5308969,15"
    assert format_money(D("0.1")) == "0,10"
    assert format_money(D("1000")) == "1000,00", "разрядов не группируем нигде"
