"""Исполнитель: поле, которое включается употреблением (ADR-028).

Проверяется главным образом то, чего быть НЕ должно: колонки у тех, кто им не
пользуется; исполнителя в слепке; выплат в документе заказчика; липкости,
пережившей ночь.
"""

from datetime import timedelta
from decimal import Decimal as D

import pytest
from openpyxl import load_workbook

from conftest import open_storage
from smeta_core import Category, PositionData, frozen_hash
from smeta_export import build_workbook
from smeta_storage import (
    STICKY_HOURS,
    create_estimate,
    history,
    performers,
    positions,
    utcnow,
)

UID = 777


@pytest.fixture
def db(tmp_path):
    """Своя база на тест: исполнитель трогает схему, и общая мешала бы."""
    _, Session = open_storage(tmp_path / "crew.db")
    with Session() as session:
        yield session


def line(name="Штукатурка", price="500", category=Category.WORK):
    return PositionData(category=category, name=name, qty=D("40"), price=D(price),
                        unit="м²", unit_spoken="квадратов")


# --- Липкость видна и протухает ---


def test_sticky_expires_after_a_night(db):
    performers.remember(db, UID, "Саня")
    assert performers.sticky(db, UID) == "Саня"

    state = performers.user_state(db, UID)
    state.performer_touched_at = utcnow() - timedelta(hours=STICKY_HOURS, minutes=1)
    db.commit()
    assert performers.sticky(db, UID) == "", "утро — не продолжение вчерашней пачки"


def test_sticky_survives_a_lunch_break(db):
    performers.remember(db, UID, "Саня")
    state = performers.user_state(db, UID)
    state.performer_touched_at = utcnow() - timedelta(hours=STICKY_HOURS - 1)
    db.commit()
    assert performers.sticky(db, UID) == "Саня"


def test_the_window_slides_on_use_not_on_setting(db):
    """Отвалившийся посреди рабочего дня хуже прилипшего."""
    performers.remember(db, UID, "Саня")
    state = performers.user_state(db, UID)
    state.performer_touched_at = utcnow() - timedelta(hours=STICKY_HOURS - 1)
    db.commit()

    performers.touch(db, UID)
    state = performers.user_state(db, UID)
    state.performer_touched_at = utcnow() - timedelta(hours=STICKY_HOURS - 1)
    db.commit()
    assert performers.sticky(db, UID) == "Саня", "употребление продлевает окно"


def test_off_clears_it_without_touching_what_is_written(db):
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line(), performer="Саня")
    db.commit()

    performers.remember(db, UID, "Саня")
    performers.forget_sticky(db, UID)

    assert performers.sticky(db, UID) == ""
    assert positions.load(db, UID, estimate.id)[0].performer == "Саня"


# --- Массовая простановка ---


def test_bulk_assignment_does_not_overwrite_what_was_named_by_hand(db):
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line("Штукатурка"), performer="Паша")
    positions.add(db, UID, estimate.id, line("Шпаклёвка"))
    positions.add(db, UID, estimate.id, line("Плитка"))
    db.commit()

    touched, skipped = performers.assign(db, UID, estimate.id, "Саня")

    assert (touched, skipped) == (2, []), "только те, у кого исполнителя не было"
    by_name = {row.name: row.performer for row in positions.load(db, UID, estimate.id)}
    assert by_name == {"Штукатурка": "Паша", "Шпаклёвка": "Саня", "Плитка": "Саня"}


def test_named_rows_are_overwritten_because_naming_them_was_the_point(db):
    estimate = create_estimate(db, UID, name="Квартира")
    row = positions.add(db, UID, estimate.id, line("Штукатурка"), performer="Паша")
    db.commit()

    assert performers.assign(db, UID, estimate.id, "Саня", ids=[row.id]) == (1, [])
    assert positions.load(db, UID, estimate.id)[0].performer == "Саня"


def test_several_names_on_one_line_are_kept_as_said(db):
    """«по 1200» на троих — это 1200 каждому, а не треть. Не делим (ADR-028)."""
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line(), performer="Саня, Паша, Олег")
    db.commit()

    row = positions.load(db, UID, estimate.id)[0]
    assert row.performer == "Саня, Паша, Олег"
    # Сумма строки от числа людей не зависит вовсе.
    assert row.qty * row.price == D("20000")


# --- Граница документа ---


def test_the_performer_is_not_part_of_the_frozen_snapshot(db):
    """Заказчик согласовывал сумму, а не бригаду: смена людей хеш не двигает."""
    estimate = create_estimate(db, UID, name="Квартира")
    row = positions.add(db, UID, estimate.id, line(), performer="Саня")
    db.commit()

    before = frozen_hash([row.to_domain()], D("0.06"), D("0.06"))
    row.performer = "Паша"
    db.commit()
    after = frozen_hash([positions.load(db, UID, estimate.id)[0].to_domain()],
                        D("0.06"), D("0.06"))

    assert before == after


def test_the_column_appears_only_when_someone_uses_it(db):
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line())
    db.commit()
    materials, works = positions.by_category(db, UID, estimate.id)
    crew_materials, crew_works = positions.performers_by_category(db, UID, estimate.id)

    sheet = load_workbook(build_workbook(
        materials, works, D("0.06"), D("0.06"),
        work_performers=crew_works, material_performers=crew_materials,
    ))["Работы"]
    assert sheet.cell(row=3, column=8).value is None, "колонки нет, пока она пуста"


def test_the_column_appears_when_it_is_filled(db):
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line(), performer="Саня")
    db.commit()
    materials, works = positions.by_category(db, UID, estimate.id)
    crew_materials, crew_works = positions.performers_by_category(db, UID, estimate.id)

    sheet = load_workbook(build_workbook(
        materials, works, D("0.06"), D("0.06"),
        work_performers=crew_works, material_performers=crew_materials,
    ))["Работы"]
    assert sheet.cell(row=3, column=8).value == "Исполнитель"
    assert sheet.cell(row=4, column=8).value == "Саня"


def test_performers_line_up_with_the_positions_they_belong_to(db):
    """Два списка, один порядок. Держится не обещанием, а этим тестом."""
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line("Штукатурка"), performer="Саня")
    positions.add(db, UID, estimate.id, line("Цемент", category=Category.MATERIAL),
                  performer="Паша")
    positions.add(db, UID, estimate.id, line("Плитка"), performer="Олег")
    db.commit()

    materials, works = positions.by_category(db, UID, estimate.id)
    crew_materials, crew_works = positions.performers_by_category(db, UID, estimate.id)

    assert [p.name for p in works] == ["Штукатурка", "Плитка"]
    assert crew_works == ["Саня", "Олег"]
    assert [p.name for p in materials] == ["Цемент"]
    assert crew_materials == ["Паша"]


# --- История ---


def test_a_rate_belongs_to_the_person_who_named_it(db):
    """100 / 150 / 250 / 250 на двоих — не разброс одной ставки, а две разные."""
    from smeta_prices import from_history

    estimate = create_estimate(db, UID, name="Квартира")
    for price, who in (("100", "Саня"), ("150", "Саня"), ("250", "Паша"), ("250", "Паша")):
        positions.add(db, UID, estimate.id, line(price=price, name=f"Работа {price}"),
                      performer=who)
    db.commit()

    points = history.lookup(db, UID, "Работа 100", "м²", "квадратов",
                            category=Category.WORK.value)
    everyone = from_history(points)
    his = from_history(points, performer="Саня")

    assert everyone is not None and his is not None
    assert his.performer == "Саня"
    assert his.low == his.high == D("100.00")


def test_an_unknown_performer_falls_back_without_claiming_the_rate_is_his(db):
    from smeta_prices import from_history

    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line(price="500"), performer="Саня")
    db.commit()

    points = history.lookup(db, UID, "Штукатурка", "м²", "квадратов",
                            category=Category.WORK.value)
    hint = from_history(points, performer="Витя")

    assert hint is not None and hint.last == D("500.00")
    assert hint.performer == "", "чужую ставку Вите не приписываем"


def test_forget_erases_the_performer_with_everything_else(db):
    """Имя живого человека стирается вместе со строкой (ADR-017, поправка)."""
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line(), performer="Саня")
    db.commit()
    history.archive(db, UID, [estimate.id])
    db.commit()

    from smeta_storage import PriceHistory
    stored = db.query(PriceHistory).filter_by(user_id=UID).all()
    assert [row.performer for row in stored] == ["Саня"]

    history.forget(db, UID)
    db.commit()
    assert db.query(PriceHistory).filter_by(user_id=UID).all() == []


def test_two_performers_on_one_day_are_two_prices_not_one(db):
    """Ключ дня включает исполнителя, иначе вторая цена исчезала бы молча."""
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line(price="150"), performer="Саня")
    positions.add(db, UID, estimate.id, line(price="250"), performer="Паша")
    db.commit()

    assert history.archive(db, UID, [estimate.id]) == 2
    db.commit()

    from smeta_storage import PriceHistory
    stored = db.query(PriceHistory).filter_by(user_id=UID).all()
    assert sorted(row.performer for row in stored) == ["Паша", "Саня"]


@pytest.mark.parametrize("tail,expected", [
    ("Саня", ("Саня", None)),
    ("Саня #41 #43", ("Саня", [41, 43])),
    ("Саня Паша", ("Саня Паша", None)),
    ("Саня 41", ("Саня", [41])),
])
def test_the_command_tail_is_read_the_way_it_is_written(tail, expected):
    from bot.handlers.crew import parse

    assert parse(tail) == expected


# --- Труд против аренды (ADR-029) ---


def test_a_concrete_mixer_gets_no_performer(db):
    """«Час или смена → это ставка человека» затащило бы аренду в человеко-дни."""
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, PositionData(
        category=Category.WORK, name="Аренда бетономешалки",
        qty=D("2"), price=D("1500"), unit="смена", unit_spoken="смены",
    ))
    positions.add(db, UID, estimate.id, PositionData(
        category=Category.WORK, name="Работа разнорабочего",
        qty=D("3"), price=D("3000"), unit="смена", unit_spoken="смены",
    ))
    db.commit()

    touched, skipped = performers.assign(db, UID, estimate.id, "Саня")

    assert touched == 1
    assert skipped == ["Аренда бетономешалки"], "пропуск назван, а не молчалив"
    by_name = {row.name: row.performer for row in positions.load(db, UID, estimate.id)}
    assert by_name == {"Аренда бетономешалки": "", "Работа разнорабочего": "Саня"}


def test_naming_the_rental_row_by_hand_does_not_help_either(db):
    """Правило одно на всех путях: поимённая форма аренду тоже не берёт."""
    estimate = create_estimate(db, UID, name="Квартира")
    row = positions.add(db, UID, estimate.id, PositionData(
        category=Category.WORK, name="Аренда перфоратора",
        qty=D("1"), price=D("800"), unit="смена", unit_spoken="смену",
    ))
    db.commit()

    assert performers.assign(db, UID, estimate.id, "Саня", ids=[row.id]) == (
        0, ["Аренда перфоратора"]
    )
    assert positions.load(db, UID, estimate.id)[0].performer == ""


def test_a_one_off_work_the_catalogue_never_heard_of_takes_a_performer(db):
    """Хвост разовых работ — большинство строк реальной сметы (ADR-029)."""
    estimate = create_estimate(db, UID, name="Квартира")
    positions.add(db, UID, estimate.id, line("Пересборка канализации"))
    db.commit()

    touched, skipped = performers.assign(db, UID, estimate.id, "Саня")
    assert (touched, skipped) == (1, [])
