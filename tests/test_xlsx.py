"""Группа E: XLSX с живыми формулами (docs/money.md §6.E)."""

from __future__ import annotations

import io
import os
import shutil
from decimal import Decimal as D

import pytest
from openpyxl import Workbook, load_workbook

from conftest import excel_round
from smeta_core import Category, PositionData, calculate_estimate
from smeta_export import FIRST_DATA_ROW, build_sheet, build_workbook

POSITIONS = [
    PositionData(Category.WORK, "Побелка", D("1.5"), D("100.10"), "м²"),
    PositionData(Category.WORK, "Стяжка", D("2.5"), D("100.10"), "м²"),
    PositionData(Category.WORK, "Копейка", D("1"), D("0.01"), "шт"),
    PositionData(Category.WORK, "Половинка", D("0.5"), D("0.01"), "шт"),   # ровно 0.005
    PositionData(Category.WORK, "Максимум", D("99999.999"), D("5000.00"), "м.п."),
]
RATE = D("6.00")
LAST_DATA_ROW = FIRST_DATA_ROW + len(POSITIONS) - 1


@pytest.fixture
def sheet():
    workbook = Workbook()
    build_sheet(workbook.active, "Работы", POSITIONS, RATE, is_work=True)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return load_workbook(buffer).active


def test_e4_row_formulas_reference_the_rate_cell_not_a_literal(sheet):
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        formula = sheet[f"G{row}"].value
        assert "$B$1" in formula, formula
        assert "1.06" not in formula
        assert "6.00" not in formula


def test_e5_totals_are_sums_over_rounded_cells(sheet):
    total_row = LAST_DATA_ROW + 1
    assert sheet[f"F{total_row}"].value == f"=SUM(F{FIRST_DATA_ROW}:F{LAST_DATA_ROW})"
    assert sheet[f"G{total_row}"].value == f"=SUM(G{FIRST_DATA_ROW}:G{LAST_DATA_ROW})"


def test_markup_row_is_a_difference(sheet):
    total_row = LAST_DATA_ROW + 1
    assert sheet[f"G{total_row + 1}"].value == f"=G{total_row}-F{total_row}"


def test_e6_quantity_and_price_cells_round_trip_exactly(sheet):
    for index, position in enumerate(POSITIONS):
        row = FIRST_DATA_ROW + index
        assert D(str(sheet[f"D{row}"].value)) == position.qty
        assert D(str(sheet[f"E{row}"].value)) == position.price


def test_unit_column_carries_the_unit(sheet):
    for index, position in enumerate(POSITIONS):
        assert sheet[f"C{FIRST_DATA_ROW + index}"].value == position.unit


def test_missing_unit_is_shown_as_a_dash():
    workbook = Workbook()
    build_sheet(
        workbook.active, "Работы",
        [PositionData(Category.WORK, "Без единицы", D("1"), D("1.00"))],
        RATE, is_work=True,
    )
    assert workbook.active[f"C{FIRST_DATA_ROW}"].value == "—"


def test_rate_cell_holds_the_rate(sheet):
    assert D(str(sheet["B1"].value)) == RATE


def test_formulas_transcribe_the_domain_exactly(sheet):
    """Считаем формулы из файла по правилам Excel и сверяем с ядром.

    Это не второй вычислитель, а проверка транскрипции: формулы обязаны быть
    именно =ROUND(D*E,2) и =ROUND(F*(1+$B$1/100),2) и давать те же копейки.
    """
    expected = calculate_estimate(POSITIONS, RATE, RATE)
    rate = float(sheet["B1"].value)

    excel_subtotal = D("0.00")
    excel_total = D("0.00")
    for index, line in enumerate(expected.lines):
        row = FIRST_DATA_ROW + index
        assert sheet[f"F{row}"].value == f"=ROUND(D{row}*E{row},2)"
        assert sheet[f"G{row}"].value == f"=ROUND(F{row}*(1+$B$1/100),2)"

        base = excel_round(float(sheet[f"D{row}"].value) * float(sheet[f"E{row}"].value))
        total = excel_round(float(base) * (1.0 + rate / 100.0))
        assert base == line.base
        assert total == line.total
        excel_subtotal += base
        excel_total += total

    assert excel_subtotal == expected.subtotal
    assert excel_total == expected.total
    assert excel_total - excel_subtotal == expected.markup


def test_empty_sheet_still_generates():
    buffer = build_workbook([], [], RATE, RATE)
    assert buffer.getbuffer().nbytes > 0


def test_workbook_has_both_sheets():
    book = load_workbook(build_workbook(POSITIONS[:1], POSITIONS[1:], RATE, D("0.00")))
    assert book.sheetnames == ["Работы", "Материалы и расходники"]
    assert D(str(book["Материалы и расходники"]["B1"].value)) == D("0.00")


def _libreoffice_missing() -> bool:
    """В CI пропуск этого теста — не «зелено», а «не проверено».

    REQUIRE_LIBREOFFICE=1 превращает отсутствие движка в падение: иначе задание
    в workflow могло бы молча ничего не проверять.
    """
    if shutil.which("soffice") is not None:
        return False
    if os.getenv("REQUIRE_LIBREOFFICE") == "1":
        pytest.fail("REQUIRE_LIBREOFFICE=1, но soffice не найден — паритет не проверен")
    return True


def _recalculate(path, tmp_path):
    """Пересчитывает книгу настоящим LibreOffice и возвращает значения ячеек."""
    import subprocess

    out = tmp_path / "out"
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(out), str(path)],
        check=True, capture_output=True, timeout=300,
    )
    return load_workbook(out / path.name, data_only=True)


@pytest.mark.skipif(_libreoffice_missing(), reason="LibreOffice не установлен")
def test_e1_libreoffice_recalculation_matches(tmp_path):
    """E1–E3: формулы, пересчитанные живым движком, дают те же копейки.

    До сих пор паритет с Excel доказывался эмулятором excel_round. Здесь его
    считает настоящий табличный процессор — на тех же «злых» данных: границы
    0.005 в обоих каскадах, копеечная строка и значение у потолка.
    """
    workbook = Workbook()
    build_sheet(workbook.active, "Работы", POSITIONS, RATE, is_work=True)
    source = tmp_path / "estimate.xlsx"
    workbook.save(source)

    sheet = _recalculate(source, tmp_path).active
    expected = calculate_estimate(POSITIONS, RATE, RATE)

    for index, line in enumerate(expected.lines):
        row = FIRST_DATA_ROW + index
        assert D(str(sheet[f"F{row}"].value)) == line.base, f"строка {row}, без наценки"
        assert D(str(sheet[f"G{row}"].value)) == line.total, f"строка {row}, с наценкой"

    total_row = LAST_DATA_ROW + 1
    assert D(str(sheet[f"F{total_row}"].value)) == expected.subtotal
    assert D(str(sheet[f"G{total_row}"].value)) == expected.total
    # Наценка в файле — разность итогов, как и в домене.
    assert D(str(sheet[f"G{total_row + 1}"].value)) == expected.markup


@pytest.mark.skipif(_libreoffice_missing(), reason="LibreOffice не установлен")
def test_e7_changing_the_rate_cell_recalculates_the_whole_sheet(tmp_path):
    """Ставка в файле — живая: правка B1 честно меняет итог, а не декорацию."""
    workbook = Workbook()
    build_sheet(workbook.active, "Работы", POSITIONS, RATE, is_work=True)
    workbook.active["B1"] = D("10.00")          # заказчик поправил ставку сам
    source = tmp_path / "reprice.xlsx"
    workbook.save(source)

    sheet = _recalculate(source, tmp_path).active
    expected = calculate_estimate(POSITIONS, D("10.00"), D("10.00"))

    total_row = LAST_DATA_ROW + 1
    assert D(str(sheet[f"G{total_row}"].value)) == expected.total
    assert D(str(sheet[f"F{total_row}"].value)) == expected.subtotal


@pytest.mark.skipif(_libreoffice_missing(), reason="LibreOffice не установлен")
def test_both_sheets_survive_a_recalculation(tmp_path):
    materials = [PositionData(Category.MATERIAL, "Гвозди", D("1000"), D("0.37"), "шт")]
    source = tmp_path / "book.xlsx"
    with open(source, "wb") as handle:
        handle.write(build_workbook(materials, POSITIONS, RATE, D("0.00")).getvalue())

    book = _recalculate(source, tmp_path)
    works_expected = calculate_estimate(POSITIONS, RATE, RATE)
    materials_expected = calculate_estimate(materials, D("0.00"), D("0.00"))

    works = book["Работы"]
    assert D(str(works[f"G{LAST_DATA_ROW + 1}"].value)) == works_expected.total

    goods = book["Материалы и расходники"]
    assert D(str(goods[f"G{FIRST_DATA_ROW}"].value)) == materials_expected.lines[0].total


def test_the_report_prints_the_unit_as_it_was_said():
    """Заказчик читает «мешков», а не канон справочника (ADR-015)."""
    said = [PositionData(Category.MATERIAL, "Цемент", D("20"), D("350.00"),
                         unit="шт", unit_spoken="мешков")]
    book = load_workbook(io.BytesIO(build_workbook(said, [], RATE, RATE).getvalue()))
    sheet = book["Материалы и расходники"]
    assert sheet[f"C{FIRST_DATA_ROW}"].value == "мешков"


def test_without_a_spoken_unit_the_canon_is_printed():
    canon_only = [PositionData(Category.MATERIAL, "Гвозди", D("100"), D("20.00"), unit="шт")]
    book = load_workbook(io.BytesIO(build_workbook(canon_only, [], RATE, RATE).getvalue()))
    assert book["Материалы и расходники"][f"C{FIRST_DATA_ROW}"].value == "шт"
